import datetime
import typing
from datetime import date

from loguru import logger
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet

from rating.models import Result, Standard, Discipline

MAX_ROW_SEARCH = 1000
DAY_ROW = 1
YEAR_ROW = 2
HORSE_NAME_ROW = 3
ATHLETE_NAME_ROW = 4
CLUB_NAME_ROW = 5
STANDARD_NAME_ROW = 1
STANDARD_NAME_COL = 4
STANDARD_DESCRIPTION_ROW = 3
STANDARD_DESCRIPTION_COL = 3


def parse_workbook(file) -> typing.List[typing.Tuple[typing.List[Result], Standard]]:
    wb = load_workbook(file)
    results = []
    for ws in wb.worksheets:
        results.append(parse_worksheet(ws))
    return results


def parse_worksheet(ws: worksheet) -> typing.Tuple[typing.List[Result], Standard]:
    row = search_first_line(ws)
    standard = get_standard(ws)
    results = []
    while not is_empty_row(ws, row):
        day = ws.cell(row=row, column=DAY_ROW).value
        year = ws.cell(row=row, column=YEAR_ROW).value
        date_ = get_date_by_russian_date(day, year)

        horse_name = ws.cell(row=row, column=HORSE_NAME_ROW).value
        athlete_name = ws.cell(row=row, column=ATHLETE_NAME_ROW).value
        club_name = ws.cell(row=row, column=CLUB_NAME_ROW).value
        results.append(Result(
            date=date_,
            horse_name=horse_name,
            athlete_name=athlete_name,
            club_name=club_name,
        ))
        logger.debug("load new result {}", results[-1])
        row += 1
    return results, standard


def get_standard(ws: worksheet) -> Standard:
    return Standard(
        name=ws.cell(
            row=STANDARD_NAME_ROW,
            column=STANDARD_NAME_COL
        ).value,
        description=ws.cell(
            row=STANDARD_DESCRIPTION_ROW,
            column=STANDARD_DESCRIPTION_COL
        ).value,
    )


def get_date_by_russian_date(day_text: str, year: int) -> date:
    day, month = day_text.split()
    month = translate_month_name(month)
    day = int(day)
    return date(year, month, day)


def translate_month_name(month_name_ru: str) -> int:
    converter = {
        'января': 1,
        'февраля': 2,
        'марта': 3,
        'апреля': 4,
        'мая': 5,
        'июня': 6,
        'июля': 7,
        'августа': 8,
        'сентября': 9,
        'октября': 10,
        'ноября': 11,
        'декабря': 12,
    }
    try:
        return converter[month_name_ru]
    except KeyError:
        raise ValueError(
            "In month mast be str with russian name of month, "
            f"found {month_name_ru}"
        )


def search_first_line(ws: worksheet) -> int:
    for row_index in range(1, MAX_ROW_SEARCH):
        if ws.cell(row=row_index, column=DAY_ROW).value == "Дата":
            return row_index + 1  # use next row after header


def is_empty_row(ws: worksheet, row: int) -> bool:
    return ws.cell(row=row, column=DAY_ROW).value is None


def save_results(results: typing.List[Result], to_discipline: Discipline, standard: Standard, using=None):
    standard.discipline = to_discipline
    standard.save(using=using)
    for result in results:
        result.fulfilled_standard = standard
    # noinspection PyUnresolvedReferences
    Result.objects.bulk_create(results)
